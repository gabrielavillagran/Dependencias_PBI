import streamlit as st
import pandas as pd
import networkx as nx
from pyvis.network import Network
import tempfile
import os
import json
import streamlit.components.v1 as components
import zipfile
import re
from pathlib import Path
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(layout="wide", page_title="Semantic Model Insights")

# --- CSS PERSONALIZADO COM ANIMAÇÕES (MELHORIA 26) ---
st.markdown("""
    <style>
        /* Sidebar responsiva - Removido CSS fixo para evitar quebra de layout nativo */
        
        /* Animações suaves nos cards */
        .stMetric { 
            background-color: #f8f9fb; 
            padding: 10px; 
            border-radius: 10px; 
            border: 1px solid #e6e9ef;
            transition: all 0.3s ease;
        }
        .stMetric:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            border-color: #5E9AE9;
        }
        
        /* Animação de fade-in */
        .element-container {
            animation: fadeIn 0.5s ease-in;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        /* Efeito hover nos botões */
        .stDownloadButton > button {
            transition: all 0.3s ease;
        }
        .stDownloadButton > button:hover {
            transform: scale(1.05);
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }
        
        /* Reduzir tamanho de fonte para simular zoom 80% */
        .main .block-container {
            font-size: 0.85rem;
        }
        
        /* Ajustar títulos */
        .main h1 { font-size: 1.8rem !important; }
        .main h2 { font-size: 1.4rem !important; }
        .main h3 { font-size: 1.1rem !important; }
        
        /* Ajustar métricas */
        .stMetric label { font-size: 0.8rem !important; }
        .stMetric [data-testid="stMetricValue"] { font-size: 1.4rem !important; }
        
        /* Ajustar sidebar - fonte menor e largura reduzida */
        [data-testid="stSidebar"] {
            width: 280px !important;
            min-width: 280px !important;
        }
        [data-testid="stSidebar"] > div:first-child {
            width: 280px !important;
        }
        .stSidebar { font-size: 0.75rem !important; }
        .stSidebar h1, .stSidebar h2, .stSidebar h3 { font-size: 0.9rem !important; }
        .stSidebar .stRadio label { font-size: 0.75rem !important; }
        .stSidebar .stMultiSelect label { font-size: 0.75rem !important; }
        .stSidebar .stTextInput label { font-size: 0.75rem !important; }
        .stSidebar .stCheckbox label { font-size: 0.75rem !important; }
        .stSidebar p { font-size: 0.75rem !important; }
        .stSidebar [data-testid="stMarkdownContainer"] { font-size: 0.75rem !important; }
        
        /* Ajustar mensagens de alerta (st.info, st.success, st.warning, st.error) */
        [data-testid="stAlert"] { font-size: 0.8rem !important; }
        [data-testid="stAlert"] p { font-size: 0.8rem !important; }
        .stAlert { font-size: 0.8rem !important; }
        div[data-baseweb="notification"] { font-size: 0.8rem !important; }
        
        /* Ajustar tabelas (dataframes) */
        [data-testid="stDataFrame"] { font-size: 0.75rem !important; width: 100% !important; }
        [data-testid="stDataFrame"] th { font-size: 0.75rem !important; }
        [data-testid="stDataFrame"] td { font-size: 0.75rem !important; }
        .stDataFrame { font-size: 0.75rem !important; }
        div[data-testid="stDataFrameResizable"] { font-size: 0.75rem !important; width: 100% !important; }
        div[data-testid="stDataFrameResizable"] > div { width: 100% !important; }
        
        /* Espaçamento entre containers de análise de impacto */
        [data-testid="stVerticalBlockBorderWrapper"] { margin-bottom: 20px !important; }
        
        /* Ajustar tamanho dos valores selecionados nos filtros (multiselect e selectbox) */
        span[data-baseweb="tag"] {
            font-size: 0.7rem !important;
            height: 20px !important;
        }
        div[data-baseweb="select"] {
            font-size: 0.8rem !important;
            min-height: 30px !important;
        }
        div[role="listbox"] {
            font-size: 0.8rem !important;
        }
        
        
        </style>
""", unsafe_allow_html=True)

st.title("Semantic Model Insights: Alta Performance & Governança DAX")

# --- FUNÇÕES AUXILIARES ---
def limpar_dax(texto):
    if pd.isnull(texto) or texto == "None":
        return ""
    return str(texto).replace("_x000D_", "").strip()

# --- FUNÇÕES DE PARSING TMDL OTIMIZADAS ---
# Compilar regex patterns uma vez (muito mais rápido)
_MEASURE_PATTERN = re.compile(r"measure\s+['\"]?([^'\"=]+)['\"]?\s*=\s*(.*)")
_BRACKET_PATTERN = re.compile(r'\[([^\]]+)\]')
_COLUMN_PATTERN = re.compile(r"'?([A-Za-z_][A-Za-z0-9_ ]*)'?\[([^\]]+)\]")
_PROPERTY_KEYWORDS = frozenset(['formatString', 'displayFolder', 'lineageTag', 
                                'annotation', 'dataCategory', 'isHidden',
                                'sourceLineageTag', 'changedProperty',
                                'formatStringDefinition'])

@st.cache_data(show_spinner=False)
def parse_tmdl_file_cached(filepath_str):
    """
    Parse a TMDL file (CACHED for speed).
    Returns a list of dicts with 'name' and 'expression' keys.
    """
    measures = []
    
    with open(filepath_str, 'r', encoding='utf-8') as f:
        lines = f.readlines()  # Ler direto em lista é mais rápido
    
    i = 0
    n_lines = len(lines)
    
    while i < n_lines:
        line = lines[i].strip()
        
        if line.startswith('measure '):
            measure_match = _MEASURE_PATTERN.match(line)
            
            if measure_match:
                measure_name = measure_match.group(1).strip()
                rest_of_line = measure_match.group(2).strip()
                
                if rest_of_line.startswith('```'):
                    # Multi-line expression
                    expression_lines = []
                    i += 1
                    while i < n_lines and lines[i].strip() != '```':
                        expression_lines.append(lines[i])
                        i += 1
                    expression = ''.join(expression_lines).strip()  # join é mais rápido que \n.join
                else:
                    # Single-line expression
                    expression_lines = [rest_of_line]
                    i += 1
                    while i < n_lines:
                        next_line = lines[i]
                        stripped = next_line.strip()
                        
                        if stripped and not next_line.startswith('\t'):
                            break
                        # Usar any() com generator é mais rápido
                        if any(stripped.startswith(kw) for kw in _PROPERTY_KEYWORDS):
                            break
                        
                        expression_lines.append(next_line)
                        i += 1
                    expression = ''.join(expression_lines).strip()
                    i -= 1
                
                measures.append((measure_name, expression))  # Tuple é mais rápido que dict
        
        i += 1
    
    return measures

def find_measure_references_fast(expression, all_measure_names_set):
    """
    Find measure references (OPTIMIZED with set lookups).
    Ignora referências em comentários (// ou --).
    """
    # Remover comentários antes de buscar referências
    lines = expression.split('\n')
    clean_lines = []
    for line in lines:
        # Remover comentários // e --
        if '//' in line:
            line = line.split('//')[0]
        if '--' in line:
            line = line.split('--')[0]
        clean_lines.append(line)
    
    clean_expression = '\n'.join(clean_lines)
    matches = _BRACKET_PATTERN.findall(clean_expression)
    return [m.strip() for m in matches if m.strip() in all_measure_names_set]

def find_column_references(expression):
    """
    Find column references in Table[Column] format.
    Returns list of tuples: (table_name, column_name)
    """
    matches = _COLUMN_PATTERN.findall(expression)
    return [(table.strip(), col.strip()) for table, col in matches]

@st.cache_data(show_spinner=False, ttl=3600)
def build_dependency_dataframe(tmdl_folder_path):
    """
    Build dependency DataFrame (CACHED and OPTIMIZED).
    """
    tmdl_files = list(Path(tmdl_folder_path).glob('*.tmdl'))
    
    if not tmdl_files:
        return None
    
    # Processar em batch - dict comprehension é mais rápido
    all_measures_list = []
    for tmdl_file in tmdl_files:
        measures = parse_tmdl_file_cached(str(tmdl_file))  # Cache hit depois da primeira vez
        all_measures_list.extend(measures)
    
    # Criar dict uma vez
    all_measures = dict(all_measures_list)
    all_measure_names = frozenset(all_measures.keys())  # frozenset é mais rápido para lookup
    
    # Criar dependências em batch
    dependencies = []
    
    # 1. Dependências de MEASURE para MEASURE
    for measure_name, expression in all_measures.items():
        # Referências a outras medidas
        for ref in find_measure_references_fast(expression, all_measure_names):
            dependencies.append({
                '[Tipo Origem]': 'MEASURE',
                '[Origem]': ref,
                '[Expressão Origem]': all_measures[ref],
                '[Tipo Destino]': 'MEASURE',
                '[Destino]': measure_name,
                '[Expressão Destino]': expression
            })
        
        # 2. Referências a colunas (Table[Column])
        for table_name, column_name in find_column_references(expression):
            col_full_name = f"{table_name}[{column_name}]"
            dependencies.append({
                '[Tipo Origem]': 'COLUMN',
                '[Origem]': col_full_name,
                '[Expressão Origem]': '',
                '[Tipo Destino]': 'MEASURE',
                '[Destino]': measure_name,
                '[Expressão Destino]': expression
            })
    
    if not dependencies:
        return None
    
    # Criar DataFrame direto é mais rápido que append
    return pd.DataFrame(dependencies)


def calcular_complexity_score(expressao, nome_medida="", medidas_dependentes=0):
    """
    Calcula Complexity  Score (0-100) com 5 dimensões baseado em SQLBI + Microsoft Learn.
    
    Dimensões:
    - D1: Funções (SUMX, RANKX, FILTER, etc)
    - D2: CALCULATE e contexto
    - D3: Estrutura (linhas, VAR, comentários)
    - D4: Dependências
    - D5: Anti-patterns
    
    Returns:
        (score, classificacao, detalhes)
    """
    import re
    
    if not expressao or expressao == "":
        return 0, "🟢 Simples", []
    
    score = 0
    detalhes = []
    
    # === D1: FUNÇÕES (Peso Alto) ===
    funcoes_peso = {
        'SUMX': 8, 'AVERAGEX': 8, 'MINX': 8, 'MAXX': 8,
        'RANKX': 12,
        'FILTER': 10,
        'ADDCOLUMNS': 10,
        'SUMMARIZE': 12, 'SUMMARIZECOLUMNS': 12,
        'GENERATE': 15,
        'EARLIER': 20,
        'PATH': 8, 'CONTAINSROW': 8
    }
    
    for func, penalty in funcoes_peso.items():
        count = expressao.upper().count(func)
        if count > 0:
            score += count * penalty
            detalhes.append(f"D1: {func} ({count}x) = +{count * penalty}")
    
    # === D2: CALCULATE E CONTEXTO ===
    calculate_count = expressao.upper().count('CALCULATE')
    if calculate_count > 0:
        score += calculate_count * 5
        detalhes.append(f"D2: CALCULATE ({calculate_count}x) = +{calculate_count * 5}")
    
    # Múltiplos filtros em CALCULATE
    calc_pattern = r'CALCULATE\s*\([^,]+,([^)]+)\)'
    for match in re.findall(calc_pattern, expressao, re.IGNORECASE):
        filters = match.count(',') + 1
        if filters > 1:
            penalty = (filters - 1) * 3
            score += penalty
            detalhes.append(f"D2: CALCULATE c/ {filters} filtros = +{penalty}")
    
    # ALL, ALLEXCEPT, REMOVEFILTERS
    context_funcs = {'ALL': 6, 'ALLEXCEPT': 6, 'REMOVEFILTERS': 6, 'KEEPFILTERS': 3}
    for func, penalty in context_funcs.items():
        count = expressao.upper().count(func)
        if count > 0:
            score += count * penalty
            detalhes.append(f"D2: {func} ({count}x) = +{count * penalty}")
    
    # === D3: ESTRUTURA ===
    linhas = expressao.count('\n') + 1
    if linhas > 20:
        # +10 para passar de 20, +5 a cada 20 linhas adicionais
        linhas_extras = linhas - 20
        blocos_extras = linhas_extras // 20
        penalty = 10 + (blocos_extras * 5)
        score += penalty
        detalhes.append(f"D3: >20 linhas ({linhas}) = +{penalty}")
    elif linhas > 10:
        score += 5
        detalhes.append(f"D3: >10 linhas ({linhas}) = +5")
    
    # Bônus: VAR
    if 'VAR' in expressao.upper():
        var_count = expressao.upper().count('VAR')
        bonus = var_count * 5
        score -= bonus
        detalhes.append(f"D3: VAR ({var_count}x) = -{bonus} (bônus)")
    
    # Bônus: Comentários
    comentarios = expressao.count('--') + expressao.count('//')
    if comentarios > 0:
        bonus = min(comentarios * 2, 10)
        score -= bonus
        detalhes.append(f"D3: Comentários ({comentarios}) = -{bonus} (bônus)")
    
    # === D4: DEPENDÊNCIAS ===
    if medidas_dependentes > 0:
        penalty = medidas_dependentes * 4
        score += penalty
        detalhes.append(f"D4: {medidas_dependentes} dependentes = +{penalty}")
    
    # === D5: ANTI-PATTERNS ===
    if re.search(r'FILTER\s*\(\s*ALL\s*\(', expressao, re.IGNORECASE):
        score += 20
        detalhes.append("D5: FILTER(ALL(Tabela)) = +20")
    
    if 'DATE' in expressao.upper() and not any(x in expressao.upper() for x in ['SAMEPERIODLASTYEAR', 'DATESYTD', 'TOTALYTD', 'DATEADD']):
        score += 8
        detalhes.append("D5: Time intelligence manual = +8")
    
    # === CLASSIFICAÇÃO ===
    final_score = min(100, max(0, score))
    
    if final_score <= 20:
        classificacao = "🟢 Simples"
    elif final_score <= 40:
        classificacao = "🟡 Moderada"
    elif final_score <= 60:
        classificacao = "🟠 Complexa"
    elif final_score <= 80:
        classificacao = "🔴 Muito Complexa"
    else:
        classificacao = "⚫ Crítica"
    
    return final_score, classificacao, detalhes


def gerar_relatorio_texto(metricas, medidas_orfas, medidas_impacto, top_complexas=None, df_structure=None):
    """Gera relatório em texto para download (MELHORIA 24)"""
    
    # Formatar lista de medidas complexas
    secao_complexidade = ""
    if top_complexas:
        lista_formatada = chr(10).join(f"  • {m['medida']} (Score: {m['score']} - {m['classificacao']})" for m in top_complexas[:10])
        secao_complexidade = f"""
🔥 TOP 10 MEDIDAS MAIS COMPLEXAS (CRÍTICAS)
────────────────────────────────────────────────────────────
{lista_formatada}
"""

    # Formatar lista de medidas por página
    secao_paginas = ""
    if df_structure is not None:
        paginas_info = []
        todas_paginas = df_structure['Página'].unique().tolist()
        for pagina in sorted(todas_paginas):
            mask = df_structure['Página'] == pagina
            medidas_na_pagina = set()
            for m_list in df_structure[mask]['Medidas'].dropna():
                # Assumindo que 'Medidas' é uma string com medidas separadas por vírgula ou similar
                # mas baseado no código anterior, parece que fazemos .str.contains
                # Se for o CSV do pbi_structure_analysis, a coluna Medidas contém as medidas do visual
                for m in m_list.split(','):
                    medidas_na_pagina.add(m.strip())
            
            if medidas_na_pagina:
                lista_medidas = chr(10).join(f"    - {m}" for m in sorted(list(medidas_na_pagina)) if m)
                paginas_info.append(f"📄 Página: {pagina}\n{lista_medidas}")
        
        if paginas_info:
            secao_paginas = f"""
📑 MEDIDAS POR PÁGINA
────────────────────────────────────────────────────────────
{chr(10).join(paginas_info)}
"""

    relatorio = f"""═══════════════════════════════════════════════════════════
            RELATÓRIO DE DEPENDÊNCIAS DAX - POWER BI
═══════════════════════════════════════════════════════════

📊 MÉTRICAS GERAIS
────────────────────────────────────────────────────────────
Objetos no Modelo: {metricas.get('objetos', 0)}
Nós no Grafo: {metricas.get('nos', 0)}
Relacionamentos: {metricas.get('relacionamentos', 0)}
Medidas para Descarte: {metricas.get('orfas', 0)}
Impacto Total: {metricas.get('impacto', 0)}
{secao_complexidade}{secao_paginas}
⚠️ SUGESTÃO DE DESCARTE SEGURO ({len(medidas_orfas)})
────────────────────────────────────────────────────────────
O que são estas Medidas?
Estas medidas foram identificadas como candidatas a descarte pois:
1. NÃO são referenciadas por nenhuma outra medida (DAX).
2. NÃO foram encontradas em nenhum visual ou página do relatório.

Lista de Medidas para Descarte:
{chr(10).join(f"  • {m}" for m in sorted(list(medidas_orfas))) if medidas_orfas else "  Nenhuma medida desnecessária encontrada!"}

📊 ANÁLISE DE IMPACTO
────────────────────────────────────────────────────────────
{chr(10).join(f"  • {m['medida']}: {m['impacto']} objetos dependentes" for m in medidas_impacto) if medidas_impacto else "  Nenhuma medida selecionada"}

═══════════════════════════════════════════════════════════
Relatório gerado automaticamente
═══════════════════════════════════════════════════════════
"""
    return relatorio

def gerar_relatorio_excel(metricas, todas_medidas_complexas, candidatas_descarte, df_st, global_dependentes_count, info_map):
    """
    Gera relatório Excel profissional com múltiplas abas formatadas.
    """
    output = BytesIO()
    wb = Workbook()
    
    # Estilos reutilizáveis
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12, name='Segoe UI')
    cell_font = Font(size=10, name='Segoe UI')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # === ABA 1: RESUMO EXECUTIVO ===
    ws_resumo = wb.active
    ws_resumo.title = "📊 Resumo Executivo"
    
    # Título principal
    ws_resumo['A1'] = "ANÁLISE DE DEPENDÊNCIAS DAX - POWER BI"
    ws_resumo['A1'].font = Font(bold=True, size=16, color="2E5090", name='Segoe UI')
    ws_resumo.merge_cells('A1:D1')
    ws_resumo['A1'].alignment = center_align
    ws_resumo.row_dimensions[1].height = 30
    
    # Métricas principais
    ws_resumo['A3'] = "MÉTRICAS GERAIS"
    ws_resumo['A3'].font = Font(bold=True, size=12, color="2E5090", name='Segoe UI')
    ws_resumo.merge_cells('A3:D3')
    
    metrics_data = [
        ['Métrica', 'Valor', 'Descrição'],
        ['Objetos no Modelo', metricas.get('objetos', 0), 'Total de medidas, colunas e tabelas'],
        ['Relacionamentos', metricas.get('relacionamentos', 0), 'Dependências DAX mapeadas'],
        ['Medidas para Descarte', metricas.get('orfas', 0), 'Não usadas em fórmulas ou visuais'],
        ['Complexidade Média', f"{round(sum(m['score'] for m in todas_medidas_complexas) / len(todas_medidas_complexas), 1) if todas_medidas_complexas else 0}/100", 'Score médio de todas as medidas']
    ]
    
    for row_idx, row_data in enumerate(metrics_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_resumo.cell(row=row_idx, column=col_idx, value=value)
            cell.font = header_font if row_idx == 4 else cell_font
            cell.fill = header_fill if row_idx == 4 else PatternFill()
            cell.border = thin_border
            cell.alignment = center_align if col_idx == 2 or row_idx == 4 else Alignment(vertical='center')
    
    # Auto-width
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20
    ws_resumo.column_dimensions['C'].width = 50
    
    # === ABA 2: RANKING DE COMPLEXIDADE ===
    ws_complex = wb.create_sheet("🔥 Complexidade")
    
    df_complex = pd.DataFrame(todas_medidas_complexas).sort_values('score', ascending=False)
    
    # Escrever cabeçalho
    headers = ['Posição', 'Medida', 'Score', 'Classificação']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_complex.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Escrever dados
    for row_idx, (idx, row) in enumerate(df_complex.iterrows(), start=2):
        ws_complex.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws_complex.cell(row=row_idx, column=2, value=row['medida'])
        ws_complex.cell(row=row_idx, column=3, value=row['score']).alignment = center_align
        # Remover emojis da classificação (apenas texto)
        classificacao_texto = row['classificacao'].split(' ')[-1] if ' ' in row['classificacao'] else row['classificacao']
        ws_complex.cell(row=row_idx, column=4, value=classificacao_texto).alignment = center_align
        
        # Formatação condicional por score
        score_val = row['score']
        if score_val >= 80:
            color = "FF4444"  # Vermelho
        elif score_val >= 60:
            color = "FF9800"  # Laranja
        elif score_val >= 40:
            color = "FFC107"  # Amarelo
        else:
            color = "4CAF50"  # Verde
        
        for col in range(1, 5):
            cell = ws_complex.cell(row=row_idx, column=col)
            cell.font = cell_font
            cell.border = thin_border
            if col == 3:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10, name='Segoe UI')
    
    ws_complex.column_dimensions['A'].width = 12
    ws_complex.column_dimensions['B'].width = 45
    ws_complex.column_dimensions['C'].width = 15
    ws_complex.column_dimensions['D'].width = 20
    
    # === ABA 3: DESCARTE SEGURO ===
    ws_trash = wb.create_sheet("🗑️ Descarte Seguro")
    
    score_map = {m['medida']: m['score'] for m in todas_medidas_complexas}
    trash_data = sorted(
        [{'Medida': m, 'Score': score_map.get(m, 0)} for m in candidatas_descarte],
        key=lambda x: x['Score'],
        reverse=True
    )
    
    headers = ['Medida', 'Complexidade', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_trash.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    for row_idx, item in enumerate(trash_data, start=2):
        ws_trash.cell(row=row_idx, column=1, value=item['Medida'])
        ws_trash.cell(row=row_idx, column=2, value=item['Score']).alignment = center_align
        ws_trash.cell(row=row_idx, column=3, value='✅ Seguro para deletar').alignment = center_align
        
        for col in range(1, 4):
            cell = ws_trash.cell(row=row_idx, column=col)
            cell.font = cell_font
            cell.border = thin_border
    
    ws_trash.column_dimensions['A'].width = 50
    ws_trash.column_dimensions['B'].width = 15
    ws_trash.column_dimensions['C'].width = 25
    
    # === ABA 4: MEDIDAS POR PÁGINA ===
    if df_st is not None:
        ws_pages = wb.create_sheet("📄 Por Página")
        
        # Processar dados
        page_data = []
        for page in sorted(df_st['Página'].unique()):
            medidas_page = set()
            for _, row in df_st[df_st['Página'] == page].iterrows():
                m_raw = str(row['Medidas'])
                if m_raw and m_raw != 'nan':
                    for m in m_raw.split(','):
                        if m.strip():
                            medidas_page.add(m.strip())
            
            if medidas_page:
                scores = [score_map.get(m, 0) for m in medidas_page]
                avg_complexity = round(sum(scores) / len(scores), 1) if scores else 0
                page_data.append({
                    'Página': page,
                    'Total Medidas': len(medidas_page),
                    'Complexidade Média': avg_complexity
                })
        
        headers = ['Página', 'Total de Medidas', 'Complexidade Média']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_pages.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        for row_idx, item in enumerate(page_data, start=2):
            ws_pages.cell(row=row_idx, column=1, value=item['Página'])
            ws_pages.cell(row=row_idx, column=2, value=item['Total Medidas']).alignment = center_align
            ws_pages.cell(row=row_idx, column=3, value=item['Complexidade Média']).alignment = center_align
            
            for col in range(1, 4):
                cell = ws_pages.cell(row=row_idx, column=col)
                cell.font = cell_font
                cell.border = thin_border
        
        ws_pages.column_dimensions['A'].width = 40
        ws_pages.column_dimensions['B'].width = 20
        ws_pages.column_dimensions['C'].width = 22
    
    # === ABA 5: TOP DEPENDÊNCIAS ===
    ws_deps = wb.create_sheet("🔗 Top Dependências")
    
    deps_data = sorted(
        [{'Medida': m, 'Dependentes': global_dependentes_count.get(m, 0)} 
         for m in info_map.keys() if info_map[m].get('tipo') == 'MEASURE'],
        key=lambda x: x['Dependentes'],
        reverse=True
    )[:50]  # Top 50
    
    headers = ['Posição', 'Medida', 'Nº de Dependentes', 'Impacto']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_deps.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    for row_idx, item in enumerate(deps_data, start=2):
        ws_deps.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws_deps.cell(row=row_idx, column=2, value=item['Medida'])
        ws_deps.cell(row=row_idx, column=3, value=item['Dependentes']).alignment = center_align
        
        # Classificar impacto
        deps_count = item['Dependentes']
        if deps_count >= 20:
            impact = "🔴 Crítico"
            color = "FF4444"
        elif deps_count >= 10:
            impact = "🟠 Alto"
            color = "FF9800"
        elif deps_count >= 5:
            impact = "🟡 Médio"
            color = "FFC107"
        else:
            impact = "🟢 Baixo"
            color = "4CAF50"
        
        cell_impact = ws_deps.cell(row=row_idx, column=4, value=impact)
        cell_impact.alignment = center_align
        cell_impact.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_impact.font = Font(bold=True, color="FFFFFF", size=10, name='Segoe UI')
        
        for col in range(1, 5):
            ws_deps.cell(row=row_idx, column=col).font = cell_font
            ws_deps.cell(row=row_idx, column=col).border = thin_border
    
    ws_deps.column_dimensions['A'].width = 12
    ws_deps.column_dimensions['B'].width = 50
    ws_deps.column_dimensions['C'].width = 20
    ws_deps.column_dimensions['D'].width = 18
    
    # Desabilitar linhas de grade em todas as abas
    for sheet in wb:
        sheet.sheet_view.showGridLines = False
    
    # Salvar
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# --- 1. SESSÃO DE INSTRUÇÕES E UPLOAD ---
with st.expander("📖 Como usar este analisador?", expanded=False):
    st.markdown("""
    ### Passo a Passo:
    1. Localize a pasta do seu **Power BI Project** (formato `.pbip`) no seu computador
    2. **Antes de compactar**, faça uma cópia da pasta do seu projeto PBIP para não perder dados importantes.
    3. Na cópia, acesse a subpasta `.SemanticModel` e **remova o arquivo `cache.abf`** (ele pode deixar o ZIP muito grande e impedir o upload).
    4. **Compacte a pasta copiada** em um arquivo ZIP:
        - Clique com o botão direito na pasta copiada
        - Selecione "Enviar para > Pasta compactada"
        - Ou use 7-Zip/WinRAR se preferir
    5. **Faça o upload do arquivo ZIP** abaixo
    6. O aplicativo irá automaticamente:
        - ✅ Extrair os arquivos TMDL (Modelo Semântico)
        - ✅ Analisar a estrutura de Páginas e Visuais (Relatório)
        - ✅ Analisar todas as medidas e suas dependências
        - ✅ Gerar o grafo de dependências interativo
    
    ### ⚡ Vantagens:
    - **Sem necessidade de DAX Query**: Não precisa mais extrair dependências manualmente do Power BI!
    - **Análise completa**: Todas as medidas são processadas automaticamente
    - **Rápido**: Upload e análise em poucos segundos
    """)
    
    st.warning("⚠️ **Importante**: O arquivo ZIP deve ter no máximo 200MB. Se o seu projeto ultrapassar esse limite, remova o arquivo `cache.abf` conforme instruções acima.")

# --- 2. CARREGAMENTO DO ZIP ---
uploaded_file = st.file_uploader(
    "📁 Envie o arquivo ZIP do seu Power BI Project (.pbip)", 
    type=["zip"],
    help="Compacte a pasta do projeto PBIP e faça upload aqui"
)

# --- FUNÇÕES DE ESTRUTURA (REPORT) ---
def extract_measures_from_query(query_obj):
    measures = set()
    def recursive_search(obj):
        if isinstance(obj, dict):
            if "Measure" in obj:
                m = obj["Measure"]
                if isinstance(m, dict) and "Property" in m: measures.add(m["Property"])
            if "Aggregation" in obj:
                agg = obj["Aggregation"]
                if isinstance(agg, dict) and "Expression" in agg: recursive_search(agg["Expression"])
            for value in obj.values(): recursive_search(value)
        elif isinstance(obj, list):
            for item in obj: recursive_search(item)
    if query_obj: recursive_search(query_obj)
    return sorted(list(measures))

def extract_visual_info(visual_path):
    try:
        with open(visual_path, 'r', encoding='utf-8') as f: visual_data = json.load(f)
        v_name = visual_data.get("name", "Unknown")
        v_type = "Unknown"
        if "visual" in visual_data and isinstance(visual_data["visual"], dict):
            v_type = visual_data["visual"].get("visualType", "Unknown")
        measures = []
        if "visual" in visual_data and isinstance(visual_data["visual"], dict):
            if "query" in visual_data["visual"]: measures = extract_measures_from_query(visual_data["visual"]["query"])
            if "objects" in visual_data["visual"]: measures.extend(extract_measures_from_query(visual_data["visual"]["objects"]))
            # Detectar medidas em formatação condicional
            if "visualContainerObjects" in visual_data["visual"]: measures.extend(extract_measures_from_query(visual_data["visual"]["visualContainerObjects"]))
            if "singleVisual" in visual_data["visual"]: measures.extend(extract_measures_from_query(visual_data["visual"]["singleVisual"]))
        return {"visual_name": v_name, "visual_type": v_type, "measures": sorted(list(set(measures)))}
    except: return None

def build_structure_dataframe(report_folder):
    pages_path = Path(report_folder) / "definition" / "pages"
    if not pages_path.exists(): return None
    results = []
    for page_dir in sorted(pages_path.iterdir()):
        if page_dir.is_dir():
            p_json = page_dir / "page.json"
            if p_json.exists():
                try:
                    with open(p_json, 'r', encoding='utf-8') as f: p_data = json.load(f)
                    p_display = p_data.get("displayName", "Unknown")
                    v_dir = page_dir / "visuals"
                    if v_dir.exists() and v_dir.is_dir():
                        for v_path in sorted(v_dir.iterdir()):
                            if v_path.is_dir():
                                v_json = v_path / "visual.json"
                                if v_json.exists():
                                    v_info = extract_visual_info(v_json)
                                    if v_info: results.append({"Página": p_display, "Visual": v_info["visual_name"], "Medidas": ", ".join(v_info["measures"])})
                    else:
                        results.append({"Página": p_display, "Visual": "Nenhum visual", "Medidas": ""})
                except: pass
    return pd.DataFrame(results) if results else None


if uploaded_file:
    # Usar session_state para armazenar o DataFrame processado
    file_key = f"{uploaded_file.name}_{uploaded_file.size}"
    
    if 'current_file_key' not in st.session_state or st.session_state.current_file_key != file_key:
        st.info("⏳ Processando arquivo ZIP...")
        
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
                    for member in zip_ref.namelist():
                        # Ignora qualquer arquivo chamado cache.abf em qualquer subpasta
                        if member.lower().endswith('cache.abf'):
                            continue
                        zip_ref.extract(member, temp_dir)
                
                # 1. Procurar Modelo Semântico (TMDL)
                tmdl_folder = None
                report_folder = None
                for root, dirs, files in os.walk(temp_dir):
                    if root.endswith(os.path.join('definition', 'tables')) or root.endswith('definition\\tables'):
                        tmdl_folder = root
                    if root.endswith('.Report') or root.endswith('.report'):
                        report_folder = root
                
                if not tmdl_folder:
                    st.error("❌ Não foi possível encontrar a pasta `.SemanticModel/definition/tables` no ZIP.")
                    st.stop()
                
                # 2. Processar TMDL
                with st.spinner("🔄 Analisando medidas e dependências..."):
                    df = build_dependency_dataframe(tmdl_folder)
                    
                    # Ler TODAS as medidas do modelo (incluindo isoladas)
                    tmdl_files = list(Path(tmdl_folder).glob('*.tmdl'))
                    todas_medidas_modelo = set()
                    for tmdl_file in tmdl_files:
                        measures = parse_tmdl_file_cached(str(tmdl_file))
                        for m in measures:
                            todas_medidas_modelo.add(m[0])  # m é tupla (name, expression)
                
                if df is None or df.empty:
                    st.error("❌ Nenhuma medida ou dependência encontrada.")
                    st.stop()
                
                # 3. Processar Estrutura (Report)
                df_st_new = None
                if report_folder:
                    with st.spinner("🔄 Analisando estrutura das páginas..."):
                        df_st_new = build_structure_dataframe(report_folder)
                
                # SALVAR NO SESSION STATE E LIMPAR CACHES
                st.session_state.current_file_key = file_key
                st.session_state.df_cached = df.copy()
                st.session_state.df_st_cached = df_st_new.copy() if df_st_new is not None else None
                st.session_state.todas_medidas_modelo = todas_medidas_modelo  # Salvar TODAS as medidas
                
                # Limpar caches de análise (forçar recalculo para novo arquivo)
                for cache_key in ['analise_global_cache', 'relatorios_global_cache', 'pages_analysis_cache']:
                    if cache_key in st.session_state:
                        del st.session_state[cache_key]
                
                df_st = st.session_state.df_st_cached
                st.success("✅ Análise concluída com sucesso!")
                
            except Exception as e:
                st.error(f"❌ Erro ao processar o arquivo: {str(e)}")
                st.stop()
    else:
        df = st.session_state.df_cached
        df_st = st.session_state.get('df_st_cached')

    col_origem, col_destino = "[Origem]", "[Destino]"
    col_tipo_origem, col_exp_origem, col_exp_destino = "[Tipo Origem]", "[Expressão Origem]", "[Expressão Destino]"

    if col_origem in df.columns and col_destino in df.columns:
        df[col_origem] = df[col_origem].astype(str).replace('nan', None)
        df[col_destino] = df[col_destino].astype(str).replace('nan', None)
        df = df.dropna(subset=[col_origem, col_destino])

        # --- NAVEGAÇÃO ---
        st.sidebar.header("Navegação")
        menu = st.sidebar.radio("Ir para:", ["Análise Global", "Análise por Medida"], index=1)
        st.sidebar.markdown("---")

        # --- 3. CÁLCULOS GLOBAIS (Pre-processamento) ---
        # Cache de mapeamento de info (leve, pode rodar sempre)
        cache_info_key = 'info_map_cache'
        if cache_info_key not in st.session_state:
            info_map = {}
            for _, row in df.iterrows():
                dest = str(row[col_destino])
                orig = str(row[col_origem])
                info_map[dest] = {"exp": limpar_dax(row[col_exp_destino]), "tipo": "MEASURE"}
                if orig not in info_map or not info_map[orig]["exp"]:
                    info_map[orig] = {"exp": limpar_dax(row[col_exp_origem]), "tipo": str(row[col_tipo_origem])}
            st.session_state[cache_info_key] = info_map
        else:
            info_map = st.session_state[cache_info_key]

        # --- CÁLCULOS PESADOS - SOMENTE PARA ANÁLISE GLOBAL (Cachear!) ---
        if menu == "Análise por Medida":
            st.sidebar.header("Filtros da Análise")
            tipos_disponiveis = sorted(df[col_tipo_origem].unique().astype(str).tolist())
            selecionar_todos = st.sidebar.checkbox("Selecionar todos os tipos", value=False)
            
            if selecionar_todos:
                tipos_selecionados = st.sidebar.multiselect("Filtrar Origens por Tipo:", options=tipos_disponiveis, default=tipos_disponiveis)
            else:
                padrão = ["MEASURE"] if "MEASURE" in tipos_disponiveis else []
                tipos_selecionados = st.sidebar.multiselect("Filtrar Origens por Tipo:", options=tipos_disponiveis, default=padrão)

            df_filtrado = df[df[col_tipo_origem].isin(tipos_selecionados)]
            todas_destinos = sorted([str(m) for m in df[col_destino].unique()])
            
            # Buscar Medida
            st.sidebar.markdown("---")
            busca_medida = st.sidebar.text_input("🔍 Buscar Medida:", "", placeholder="Digite para filtrar...")
            medidas_filtradas = [m for m in todas_destinos if busca_medida.lower() in m.lower()] if busca_medida else todas_destinos
            
            medidas_selecionadas = st.sidebar.multiselect(
                "Selecione as Medidas Destino:", 
                options=medidas_filtradas, 
                default=[]
            )
            
            # Opções do Grafo
            st.sidebar.markdown("---")
            st.sidebar.subheader("Configurações do Grafo")
            direcao_grafo = st.sidebar.radio(
                "Escolha o que visualizar:",
                options=[
                    "⬇️ Dependências (do que a medida depende)",
                    "⬆️ Dependentes (quem depende da medida)"
                ],
                index=0
            )
            
            modo_visualizacao = st.sidebar.radio(
                "Modo de Visualização:",
                options=[
                    "Grafo Completo (todos os níveis)",
                    "Grafo Expansível (clique DUPLO para expandir)"
                ],
                index=0
            )
            
            export_placeholder = st.sidebar.container()
        else:
            medidas_selecionadas = []
            df_filtrado = df.copy()
            direcao_grafo = "⬇️"
            modo_visualizacao = "Completo"

        # === 4. ANÁLISE GLOBAL ===
        if menu == "Análise Global":
            # --- CALCULAR COMPLEXIDADE E DEPENDÊNCIAS (CACHE) ---
            cache_complexity_key = 'complexity_cache'
            if cache_complexity_key not in st.session_state:
                global_dependentes_count = df[col_destino].value_counts().to_dict()
                todas_medidas_complexas = []
                for nome_medida, info in info_map.items():
                    if info.get("tipo") == "MEASURE":
                        exp = info.get("exp", "")
                        n_dependentes = global_dependentes_count.get(nome_medida, 0)
                        score, classificacao, _ = calcular_complexity_score(exp, nome_medida, n_dependentes)
                        todas_medidas_complexas.append({
                            'medida': nome_medida, 
                            'score': score, 
                            'classificacao': classificacao
                        })
                st.session_state[cache_complexity_key] = {
                    'global_dependentes_count': global_dependentes_count,
                    'todas_medidas_complexas': todas_medidas_complexas
                }
            
            global_dependentes_count = st.session_state[cache_complexity_key]['global_dependentes_count']
            todas_medidas_complexas = st.session_state[cache_complexity_key]['todas_medidas_complexas']
            
            # Métricas Gerais em Cards
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Objetos no Modelo", len(info_map), help="Total de Tabelas, Colunas e Medidas encontradas nos arquivos TMDL do projeto.")
            m2.metric("Relacionamentos Total", len(df), help="Total de conexões diretas entre medidas (dependências DAX).")
            # --- CÁLCULO DE DESCARTE SEGURO (PARA O DASHBOARD) - CACHE ---
            cache_key = 'analise_global_cache'
            if cache_key not in st.session_state:
                # Pegar TODAS as medidas do modelo (incluindo isoladas sem dependências)
                todas_as_medidas = st.session_state.get('todas_medidas_modelo', set())
                
                # Medidas que SÃO USADAS por outras medidas (aparecem na coluna Destino)
                medidas_usadas_destino = set(d for d in df[col_destino].unique() if d in todas_as_medidas)
                
                # Medidas em visuais
                medidas_em_visuais_global = set()
                if df_st is not None:
                    for _, r in df_st.iterrows():
                        m_raw = str(r['Medidas'])
                        if m_raw and m_raw != 'nan' and m_raw != 'None':
                            for m_ctx in [x.strip() for x in m_raw.split(',') if x.strip()]:
                                medidas_em_visuais_global.add(m_ctx)
                
                # Candidatas = medidas que NÃO são usadas por outras E NÃO estão em visuais
                candidatas_descarte_global = todas_as_medidas - medidas_usadas_destino - medidas_em_visuais_global
                
                # Construir grafo completo (PESADO - cachear!)
                G_full = nx.from_pandas_edgelist(df, col_origem, col_destino, create_using=nx.DiGraph())
                
                # Top 10 mais impactantes (MUITO PESADO - cachear!)
                top_impacto = []
                if not df.empty:
                    for n in G_full.nodes():
                        if info_map.get(n, {}).get('tipo') == 'MEASURE':
                            top_impacto.append({'medida': n, 'impacto': len(nx.descendants(G_full, n))})
                    top_impacto = sorted(top_impacto, key=lambda x: x['impacto'], reverse=True)[:10]
                
                # Armazenar no cache
                st.session_state[cache_key] = {
                    'candidatas_descarte': candidatas_descarte_global,
                    'G_full': G_full,
                    'top_impacto': top_impacto,
                    'medidas_em_visuais': medidas_em_visuais_global
                }
            
            # Recuperar do cache
            candidatas_descarte_global = st.session_state[cache_key]['candidatas_descarte']
            G_full = st.session_state[cache_key]['G_full']
            top_impacto = st.session_state[cache_key]['top_impacto']
            medidas_em_visuais_global = st.session_state[cache_key]['medidas_em_visuais']
            
            m3.metric("Descarte Seguro", len(candidatas_descarte_global), help="Medidas que NÃO são usadas em fórmulas DAX e NÃO aparecem em nenhum visual do relatório. Candidatas seguras para exclusão.")
            
            # --- MÉTRICAS PARA RELATÓRIOS ---
            metr_exp = {
                'objetos': len(info_map), 
                'nos': len(G_full.nodes()), 
                'relacionamentos': len(df), 
                'orfas': len(candidatas_descarte_global),
                'impacto': 0 
            }

            # Cache de relatórios (só gerar quando solicitado via download)
            relatorio_cache_key = 'relatorios_global_cache'
            if relatorio_cache_key not in st.session_state:
                relatorio_txt = gerar_relatorio_texto(
                    metr_exp, 
                    candidatas_descarte_global, 
                    top_impacto, 
                    sorted(todas_medidas_complexas, key=lambda x: x['score'], reverse=True), 
                    df_st
                )
                excel_bytes = gerar_relatorio_excel(
                    metr_exp,
                    todas_medidas_complexas,
                    candidatas_descarte_global,
                    df_st,
                    global_dependentes_count,
                    info_map
                )
                st.session_state[relatorio_cache_key] = {
                    'txt': relatorio_txt,
                    'excel': excel_bytes
                }
            
            st.sidebar.download_button(
                "📄 Baixar Relatório Completo (TXT)", 
                st.session_state[relatorio_cache_key]['txt'], 
                "relatorio_global.txt", 
                "text/plain", 
                use_container_width=True
            )
            st.sidebar.download_button(
                "📊 Baixar Relatório Excel (Formatado)", 
                st.session_state[relatorio_cache_key]['excel'], 
                "relatorio_completo.xlsx", 
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
            score_geral = round(sum(m['score'] for m in todas_medidas_complexas) / len(todas_medidas_complexas), 1) if todas_medidas_complexas else 0
            m4.metric("Complexidade DAX Média", f"{score_geral}/100", help="Média do Score D1-D5 de todas as medidas DAX. Quanto menor, mais performático e legível é o seu modelo.")
            
            st.markdown("---")

            # --- MENSAGEM DE INTRODUÇÃO ---
            # Calcular informações dinâmicas
            total_medidas = len([m for m, info in info_map.items() if info.get("tipo") == "MEASURE"])
            total_paginas = len(df_st['Página'].unique()) if df_st is not None else 0
            
            # Classificar criticidade geral baseada no score médio
            if score_geral >= 60:
                criticidade = "🔴 Alta"
                msg_criticidade = "seu modelo apresenta alta complexidade nas medidas DAX"
            elif score_geral >= 40:
                criticidade = "🟠 Moderada"
                msg_criticidade = "seu modelo apresenta complexidade moderada nas medidas DAX"
            elif score_geral >= 20:
                criticidade = "🟡 Baixa"
                msg_criticidade = "seu modelo apresenta baixa complexidade nas medidas DAX"
            else:
                criticidade = "🟢 Muito Baixa"
                msg_criticidade = "seu modelo apresenta medidas DAX simples e bem otimizadas"
            
            # Usar markdown com tamanho de fonte consistente
            st.markdown(f"""
<div style="padding: 1rem; background-color: #d1ecf1; border-left: 4px solid #0c5460; border-radius: 0.25rem; font-size: 0.9rem; line-height: 1.6;">
    
📊 <strong>Olá! Bem-vindo à Análise Global do seu modelo semântico.</strong>

<strong>Resumo do seu relatório:</strong>
<ul style="margin-top: 0.5rem; margin-bottom: 0.5rem;">
<li>📄 <strong>{total_paginas}</strong> páginas identificadas no relatório</li>
<li>📐 <strong>{total_medidas}</strong> medidas DAX mapeadas</li>
<li>🎯 <strong>Criticidade DAX Média:</strong> {criticidade} (Score: {score_geral}/100)</li>
</ul>

<em>{msg_criticidade.capitalize()}</em>. Use este relatório para auxiliar nas análises de <strong>governança</strong>, <strong>otimização de performance</strong> e <strong>manutenção</strong> do seu modelo Power BI.

👇 <strong>Abaixo você encontrará detalhamentos por página, rankings de complexidade e sugestões de descarte seguro.</strong>

</div>
            """, unsafe_allow_html=True)


            
            st.markdown("---")

            # --- GRÁFICO DE BARRAS: MEDIDAS POR PÁGINA (Reposicionado para o topo) - CACHE ---
            if df_st is not None:
                try:
                    cache_pages_key = 'pages_analysis_cache'
                    if cache_pages_key not in st.session_state:
                        d_mat_raw = []
                        for _, r in df_st.iterrows():
                            p = r['Página']
                            meds_raw = str(r['Medidas'])
                            if meds_raw and meds_raw != 'nan':
                                for m_ctx in [x.strip() for x in meds_raw.split(',') if x.strip()]:
                                    d_mat_raw.append({"Página": p, "Medida": m_ctx})
                        
                        if d_mat_raw:
                            df_mat_full = pd.DataFrame(d_mat_raw).drop_duplicates()
                            df_count = df_mat_full.groupby('Página')['Medida'].nunique().reset_index()
                            df_count.columns = ['Página', 'Total de Medidas']
                            
                            comp_map = {m['medida']: m['score'] for m in todas_medidas_complexas}
                            page_stats = []
                            for pg in df_count['Página']:
                                meds = df_mat_full[df_mat_full['Página'] == pg]['Medida'].unique()
                                scores = [comp_map.get(m, 0) for m in meds]
                                avg_c = round(sum(scores)/len(scores), 1) if scores else 0
                                m_complex = max(meds, key=lambda x: comp_map.get(x, 0)) if len(meds)>0 else "-"
                                m_dep = max(meds, key=lambda x: global_dependentes_count.get(x, 0)) if len(meds)>0 else "-"
                                n_dep = global_dependentes_count.get(m_dep, 0)
                                page_stats.append({
                                    'Página': pg, 
                                    'Complexidade Média': avg_c, 
                                    'Top Complexidade': f"{m_complex} (Score: {comp_map.get(m_complex, 0)})", 
                                    'Top Dependências': f"{m_dep} ({n_dep} dependentes)"
                                })
                            
                            st.session_state[cache_pages_key] = {
                                'df_count': df_count,
                                'page_stats': page_stats
                            }
                        else:
                            st.session_state[cache_pages_key] = None
                    
                    # Recuperar do cache e renderizar
                    if st.session_state[cache_pages_key]:
                        df_count = st.session_state[cache_pages_key]['df_count']
                        page_stats = st.session_state[cache_pages_key]['page_stats']
                        
                        df_count = df_count.merge(pd.DataFrame(page_stats), on='Página').sort_values('Total de Medidas', ascending=True)
                        
                        st.markdown("#### Distribuição de Medidas por Página")
                        fig = px.bar(df_count, y='Página', x='Total de Medidas', orientation='h', text='Total de Medidas', color='Total de Medidas',
                                    color_continuous_scale=['#E3F2FD', '#2196F3', '#0D47A1'], 
                                    labels={'Total de Medidas': 'Nº de Medidas', 'Top Complexidade': 'Mais Complexa', 'Top Dependências': 'Mais Reutilizada'},
                                    hover_data={'Página': True, 'Total de Medidas': True, 'Complexidade Média': True, 'Top Complexidade': True, 'Top Dependências': True})
                        
                        fig.update_layout(
                            xaxis_title="Quantidade de Medidas Únicas", 
                            yaxis_title="", 
                            coloraxis_showscale=False, 
                            margin=dict(l=10, r=40, t=30, b=10),
                            height=max(400, len(df_count) * 45), 
                            template="plotly_white", 
                            hovermode="closest" # Alterado de 'y unified' para 'closest' para limpar a visão
                        )
                        fig.update_traces(texttemplate='%{text}', textposition='outside', marker_line_color='rgb(8,48,107)', marker_line_width=1, opacity=0.9)
                        fig.update_yaxes(tickfont=dict(size=12, color='#111111', family='Segoe UI Semibold'), gridcolor='#F0F2F6', automargin=True, zeroline=False)
                        fig.update_xaxes(tickfont=dict(size=12, color='#111111', family='Segoe UI Semibold'), gridcolor='#F0F2F6', zeroline=False)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        st.info("""
                        💡 **Sugestões de Otimização:**
                        *   **Páginas com poucas medidas:** Melhores para iniciar a otimização (baixo risco).
                        *   **Páginas com muitas medidas:** Considere simplificar ou dividir para ganhar performance.
                        """)
                        st.markdown("---")
                except: pass
            
            # Rankings
            st.markdown("##### Ranking de Complexidade")
            st.caption("""
            O **Score de Complexidade** (0-100) é calculado com base em 5 dimensões técnicas:
            **D1: Funções de Iteração** (SUMX, FILTER, etc) | **D2: Contexto** (CALCULATE, ALL) | 
            **D3: Estrutura** (Linhas, VARs) | **D4: Dependências** (Impacto no modelo) | 
            **D5: Anti-patterns** (Boas práticas).
            """)
            df_rk = pd.DataFrame(todas_medidas_complexas).sort_values(by="score", ascending=False)
            st.dataframe(
                df_rk[['medida', 'score', 'classificacao']], 
                hide_index=True, 
                use_container_width=True, 
                height=400,
                column_config={
                    "score": st.column_config.ProgressColumn(
                        "Complexidade",
                        help="Score de 0 a 100",
                        format="%f",
                        min_value=0,
                        max_value=100,
                        color="blue"
                    )
                }
            )
            
            with st.expander("ℹ️ Entenda o Cálculo do Score (Tabela de Referência)"):
                st.markdown("""
                O Score é a soma de penalidades (pontos positivos) e bônus (pontos negativos) baseados nas melhores práticas da Microsoft e SQLBI.
                """)
                
                c1, c2, c3 = st.columns(3)
                
                with c1:
                    st.markdown("**D1: Funções de Custo (Iteradores)**")
                    st.dataframe(pd.DataFrame([
                        {"Item": "EARLIER", "Pts": "+20"},
                        {"Item": "GENERATE", "Pts": "+15"},
                        {"Item": "RANKX, SUMMARIZE", "Pts": "+12"},
                        {"Item": "FILTER, ADDCOLUMNS", "Pts": "+10"},
                        {"Item": "SUMX, AVERAGEX...", "Pts": "+8"},
                    ]), hide_index=True, use_container_width=True)
                
                with c2:
                    st.markdown("**D2: Contexto (CALCULATE)**")
                    st.dataframe(pd.DataFrame([
                        {"Item": "CALCULATE", "Pts": "+5 (+3/filtro)"},
                        {"Item": "ALL, ALLEXCEPT...", "Pts": "+6"},
                        {"Item": "KEEPFILTERS", "Pts": "+3"},
                        {"Item": "REMOVEFILTERS", "Pts": "+6"},
                    ]), hide_index=True, use_container_width=True)

                with c3:
                    st.markdown("**D3/D4/D5: Estrutura & Boas Práticas**")
                    st.dataframe(pd.DataFrame([
                        {"Item": "D3: > 20 Linhas", "Pts": "+10, +5 a cada 20 linhas"},
                        {"Item": "D3: Uso de VAR", "Pts": "-5 (Bônus)"},
                        {"Item": "D4: Dependências", "Pts": "+4 por dep."},
                        {"Item": "D5: FILTER(ALL...)", "Pts": "+20 (Crítico)"},
                        {"Item": "D5: Data Manual", "Pts": "+8"},
                    ]), hide_index=True, use_container_width=True)
                
                st.markdown("---")
                st.markdown("**📊 Classificação de Criticidade por Score**")
                st.dataframe(pd.DataFrame([
                    {"Score": "0 - 20", "Classificação": "🟢 Muito Baixa", "Descrição": "Medida simples e otimizada"},
                    {"Score": "21 - 40", "Classificação": "🟡 Baixa", "Descrição": "Medida com complexidade aceitável"},
                    {"Score": "41 - 60", "Classificação": "🟠 Moderada", "Descrição": "Requer atenção e revisão"},
                    {"Score": "61 - 80", "Classificação": "🔴 Alta", "Descrição": "Complexidade elevada - revisar urgente"},
                    {"Score": "81 - 100", "Classificação": "🔴 Crítica", "Descrição": "Extremamente complexa - refatorar prioritário"},
                ]), hide_index=True, use_container_width=True)
            st.markdown("---")
            st.markdown("##### Mais Dependentes")
            G_gl = nx.DiGraph()
            G_gl.add_edges_from([(row[col_destino], row[col_origem]) for _, row in df.iterrows()])
            l_dp = []
            for mm in info_map:
                if info_map[mm].get("tipo") == "MEASURE":
                    l_dp.append({"Medida": mm, "Dependentes": len(nx.ancestors(G_gl, mm)) if mm in G_gl else 0})
            df_dp = pd.DataFrame(l_dp).sort_values(by="Dependentes", ascending=False)
            st.dataframe(
                df_dp, 
                hide_index=True, 
                use_container_width=True, 
                height=400,
                column_config={
                    "Dependentes": st.column_config.ProgressColumn(
                        "Dependentes",
                        help="Número de medidas que dependem desta",
                        format="%d",
                        min_value=0,
                        max_value=int(df_dp['Dependentes'].max()) if not df_dp.empty else 10,
                        color="blue"
                    )
                }
            )

            # --- NOVO: SUGESTÃO DE DESCARTE SEGURO ---
            st.markdown("---")
            st.markdown("##### 🧹 Sugestão de Descarte Seguro")
            
            # Usar o cálculo global já feito (não recalcular)
            lista_descarte = sorted(list(candidatas_descarte_global))
            
            if lista_descarte:
                st.warning(f"💡 Encontramos **{len(lista_descarte)}** medidas que parecem não ter uso no relatório ou no modelo.")
                st.markdown("""
                Estas medidas são consideradas **seguras para descarte** porque:
                1. Não são referenciadas por nenhuma outra medida (DAX).
                2. Não foram encontradas em nenhum visual ou página do relatório.
                """)
                
                # Criar um dataframe para mostrar as candidatas com sua complexidade
                score_map = {m['medida']: m['score'] for m in todas_medidas_complexas}
                df_trash = pd.DataFrame([
                    {
                        "Medida": m, 
                        "Complexidade": score_map.get(m, 0)
                    } for m in lista_descarte
                ]).sort_values(by="Complexidade", ascending=False)
                
                st.dataframe(
                    df_trash, 
                    hide_index=True, 
                    use_container_width=True,
                    column_config={
                        "Complexidade": st.column_config.ProgressColumn(
                            "Complexidade",
                            min_value=0, max_value=100, format="%d", color="orange"
                        )
                    }
                )
            else:
                st.success("✅ **Nenhuma medida desnecessária encontrada!** Todas as suas medidas estão sendo utilizadas em fórmulas ou visuais.")

            # Detalhamento por Página (Tabela Solicitada)
            if df_st is not None:
                st.markdown("---")
                st.subheader("Detalhamento: Medidas por Página e Visual")
                try:
                    d_mat = []
                    for _, r in df_st.iterrows():
                        p = r['Página']
                        meds_raw = str(r['Medidas'])
                        if meds_raw and meds_raw != 'nan':
                            for m_ctx in [x.strip() for x in meds_raw.split(',') if x.strip()]:
                                d_mat.append({"Página": p, "Medida": m_ctx})
                    
                    if d_mat:
                        df_mat = pd.DataFrame(d_mat).drop_duplicates().sort_values(['Página', 'Medida'])
                        
                        f1, f2 = st.columns(2)
                        p_opts = sorted(df_mat['Página'].unique())
                        m_opts = sorted(df_mat['Medida'].unique())
                        
                        p_sel = f1.multiselect("Filtrar Página:", p_opts, key="g_pg")
                        m_sel = f2.multiselect("Filtrar Medida:", m_opts, key="g_md")
                        
                        if p_sel: df_mat = df_mat[df_mat['Página'].isin(p_sel)]
                        if m_sel: df_mat = df_mat[df_mat['Medida'].isin(m_sel)]
                        
                        st.dataframe(df_mat, hide_index=True, use_container_width=True, height=500)

                except Exception as e:
                    st.warning(f"Erro ao processar estrutura: {e}")
            else:
                st.info("Arquivo 'pbi_structure_analysis.csv' não encontrado para detalhamento de páginas.")

        # === 5. ANÁLISE POR MEDIDA ===
        elif menu == "Análise por Medida":
            # --- CALCULAR COMPLEXIDADE APENAS SE NECESSÁRIO (para métricas) ---
            cache_complexity_key = 'complexity_cache'
            if cache_complexity_key not in st.session_state:
                global_dependentes_count = df[col_destino].value_counts().to_dict()
                todas_medidas_complexas = []
                for nome_medida, info in info_map.items():
                    if info.get("tipo") == "MEASURE":
                        exp = info.get("exp", "")
                        n_dependentes = global_dependentes_count.get(nome_medida, 0)
                        score, classificacao, _ = calcular_complexity_score(exp, nome_medida, n_dependentes)
                        todas_medidas_complexas.append({
                            'medida': nome_medida, 
                            'score': score, 
                            'classificacao': classificacao
                        })
                st.session_state[cache_complexity_key] = {
                    'global_dependentes_count': global_dependentes_count,
                    'todas_medidas_complexas': todas_medidas_complexas
                }
            
            todas_medidas_complexas = st.session_state[cache_complexity_key]['todas_medidas_complexas']
            
            if not medidas_selecionadas:
                st.info("👈 Selecione uma ou mais Medidas na barra lateral para detalhar dependências e impacto.")
            else:
                # 1. Parâmetros de Visualização
                modo_dependencias = "⬇️" in direcao_grafo
                modo_dependentes = "⬆️" in direcao_grafo
                modo_expansivel_val = "Expansível" in modo_visualizacao
                
                # 2. Construção do Grafo (DFS/BFS)
                arestas, visitados = [], set()
                if modo_expansivel_val:
                    for raiz in medidas_selecionadas:
                        visitados.add(raiz)
                        if modo_dependencias:
                            for f in df_filtrado[df_filtrado[col_destino] == raiz][col_origem].tolist(): arestas.append((raiz, f))
                        if modo_dependentes:
                            for p in df_filtrado[df_filtrado[col_origem] == raiz][col_destino].tolist(): arestas.append((p, raiz))
                else:
                    fila = list(medidas_selecionadas)
                    while fila:
                        at = fila.pop(0)
                        if at not in visitados:
                            visitados.add(at)
                            if modo_dependencias:
                                for f in df_filtrado[df_filtrado[col_destino] == at][col_origem].tolist():
                                    arestas.append((at, f))
                                    if f not in visitados: fila.append(f)
                            if modo_dependentes:
                                for p in df_filtrado[df_filtrado[col_origem] == at][col_destino].tolist():
                                    arestas.append((p, at))
                                    if p not in visitados: fila.append(p)

                G = nx.DiGraph()
                G.add_edges_from(arestas)
                for node in medidas_selecionadas: # Garantir raízes sem arestas apareçam
                    if node not in G: G.add_node(node)

                # 3. Métricas de Contexto
                scores_s = [m['score'] for m in todas_medidas_complexas if m['medida'] in medidas_selecionadas]
                avg_s = round(sum(scores_s) / len(scores_s), 1) if scores_s else 0
                
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("📌 Nós no Grafo", len(G.nodes()))
                c2.metric("🔗 Relacionamentos", len(arestas))
                orig_f = set(df_filtrado[col_origem].unique())
                dest_f = set(df_filtrado[col_destino].unique())
                
                # Calcular em quantas páginas as medidas selecionadas aparecem
                paginas_em_uso = set()
                if df_st is not None:
                    for _, row in df_st.iterrows():
                        meds_page_dirty = str(row['Medidas'])
                        if meds_page_dirty and meds_page_dirty != 'nan':
                            meds_in_page = [x.strip() for x in meds_page_dirty.split(',') if x.strip()]
                            # Verifica se ALGUMA das medidas selecionadas está nesta página
                            if any(m in meds_in_page for m in medidas_selecionadas):
                                paginas_em_uso.add(row['Página'])

                c3.metric("📄 Páginas em Uso", len(paginas_em_uso))
                c4.metric("📊 Score Médio DAX", f"{avg_s}/100")

                # 4. Preparação PyVis
                cores_map = {"MEASURE": "#88B995", "COLUMN": "#5E9AE9", "CALC_COLUMN": "#BBBBBB", "TABLE": "#F4A460", "CALC_TABLE": "#BBBBBB", "UNKNOWN": "#CCCCCC"}
                icones_map = {"MEASURE": "📊", "COLUMN": "📋", "CALC_COLUMN": "🔢", "TABLE": "📁", "CALC_TABLE": "🧮", "UNKNOWN": "❓"}
                
                net = Network(height="600px", width="100%", directed=True, bgcolor="#ffffff")
                nos_exp = set()
                if modo_expansivel_val:
                    for node in G.nodes():
                        targets = df_filtrado[df_filtrado[col_destino if modo_dependencias else col_origem] == node][col_origem if modo_dependencias else col_destino].tolist()
                        if set(targets) - set(G.nodes()): nos_exp.add(node)

                for node in G.nodes():
                    t = info_map.get(node, {}).get("tipo", "UNKNOWN")
                    ic, cr = icones_map.get(t, "❓"), cores_map.get(t, "#CCCCCC")
                    is_e = node in nos_exp
                    net.add_node(node, label=f"{ic} {node}{' ⊕' if is_e else ''}", color=cr, shape="box", font={"face": "Segoe UI", "size": 14, "bold": is_e}, borderWidth=3 if is_e else 1)
                
                for u_n, v_n in G.edges(): net.add_edge(u_n, v_n, color="#CCCCCC", width=1)
                net.set_options('{"physics":{"enabled":false}, "layout":{"hierarchical":{"enabled":true, "direction":"UD", "sortMethod":"directed", "nodeSpacing":300}}}')
                
                # 5. Renderização Grafo
                tmp_p = os.path.join(tempfile.gettempdir(), "graph_pbi.html")
                net.save_graph(tmp_p)
                with open(tmp_p, 'r', encoding='utf-8') as f: h_base = f.read()
                
                all_ids = set(df_filtrado[col_destino].unique()) | set(df_filtrado[col_origem].unique())
                d_js = {}
                for n_id in all_ids:
                    targets = df_filtrado[df_filtrado[col_destino if modo_dependencias else col_origem] == n_id][col_origem if modo_dependencias else col_destino].tolist()
                    d_js[str(n_id)] = {'filhos': [str(x) for x in targets], 'tipos': [info_map.get(str(x), {}).get("tipo", "UNKNOWN") for x in targets]}

                # Adicionar painel e estilos ANTES do </body>
                painel_html = """
                <div id="dax-panel" style="position:fixed; top:20px; right:20px; width:500px; max-height:85vh; background:#ffffff; border-radius:12px; padding:20px; overflow-y:auto; z-index:99999; display:none; box-shadow:0 4px 16px rgba(0,0,0,0.15); border:1px solid #e6e9ef; font-family: sans-serif;">
                    <button onclick="document.getElementById('dax-panel').style.display='none'" style="position:absolute; top:15px; right:15px; cursor:pointer; background:none; border:none; font-size:24px; color:#999; padding:0; width:30px; height:30px; line-height:30px; transition:color 0.2s;">&times;</button>
                    <div id="p-title" style="font-weight:bold; color:#1f77b4; margin-bottom:12px; font-size:16px; padding-right:30px;"></div>
                    <div id="p-exp" style="background:#282c34; padding:16px; border-radius:8px; overflow-x:auto; font-family:'Consolas', 'Monaco', 'Courier New', monospace; font-size:13px; line-height:1.3; color:#abb2bf; white-space:pre-wrap; tab-size:4;"></div>
                </div>
                <button id="reset-view-btn" onclick="resetGraphView()" style="position:fixed; bottom:20px; right:20px; z-index:99998; background:linear-gradient(135deg, #5E9AE9 0%, #2E5090 100%); color:white; border:none; border-radius:8px; padding:10px 20px; font-size:13px; font-weight:600; font-family:'Segoe UI', sans-serif; cursor:pointer; box-shadow:0 2px 8px rgba(0,0,0,0.2); transition:all 0.3s ease;">
                    🔄 Resetar Zoom
                </button>
                """
                
                estilos_css = """
                <style>
                    #dax-panel button:hover { color: #ff4444 !important; }
                    .dax-keyword { color: #c678dd; font-weight: bold; }
                    .dax-function { color: #61afef; font-weight: bold; }
                    .dax-string { color: #98c379; }
                    .dax-comment { color: #5c6370; font-style: italic; }
                    .dax-number { color: #d19a66; }
                    .dax-operator { color: #56b6c2; }
                    .dax-variable { color: #e5c07b; font-weight: bold; }
                    .dax-table { color: #e06c75; }
                    #reset-view-btn:hover { transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.3); }
                    #reset-view-btn:active { transform: translateY(0px); }
                </style>
                """
                
                script_js = f"""
                <script>
                    console.log('[DAX Viewer] Inicializando...');
                    
                    var infoData = {json.dumps(info_map)};
                    var depsMap = {json.dumps(d_js)};
                    var modoExp = {"true" if modo_expansivel_val else "false"};
                    var coresMap = {json.dumps(cores_map)};
                    var iconesMap = {json.dumps(icones_map)};

                    console.log('[DAX Viewer] Dados carregados:', Object.keys(infoData).length, 'medidas');
                    console.log('[DAX Viewer] Painel elemento:', document.getElementById('dax-panel'));

                    // Função de syntax highlighting simplificada
                    function highlightDAX(code) {{
                        if (!code || code === 'Sem DAX') return code;
                        
                        var esc = code.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
                        
                        // Comentários (PRIMEIRO - para proteger strings e código dentro de comentários)
                        esc = esc.replace(/(--[^\\n]*)/g, '###COMMENT_START###$1###COMMENT_END###');
                        esc = esc.replace(/(\/\/[^\\n]*)/g, '###COMMENT_START###$1###COMMENT_END###');
                        
                        // Strings (SEGUNDO - para proteger conteúdo de strings)
                        esc = esc.replace(/"([^"]*)"/g, '###STRING_START###"$1"###STRING_END###');
                        
                        // Números
                        esc = esc.replace(/\\b(\\d+(?:\\.\\d+)?)\\b/g, '###NUMBER_START###$1###NUMBER_END###');
                        
                        // Operadores (ANTES de keywords para não interferir com atributos HTML)
                        esc = esc.replace(/(&&)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(\\|\\|)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(&lt;=)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(&gt;=)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(&lt;&gt;)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(&lt;)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(&gt;)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(\\+)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(-)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(\\*)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(\\/)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        esc = esc.replace(/(=)/g, '###OPERATOR_START###$1###OPERATOR_END###');
                        
                        // Keywords
                        var kw = ['VAR', 'RETURN', 'IF', 'THEN', 'ELSE', 'SWITCH', 'TRUE', 'FALSE', 'IN', 'NOT', 'AND', 'OR', 'BLANK'];
                        kw.forEach(function(k) {{
                            var re = new RegExp('\\\\b(' + k + ')\\\\b', 'gi');
                            esc = esc.replace(re, '###KEYWORD_START###$1###KEYWORD_END###');
                        }});
                        
                        // Funções
                        var fn = ['CALCULATE', 'CALCULATETABLE', 'FILTER', 'ALL', 'ALLEXCEPT', 'ALLSELECTED',
                            'SUM', 'SUMX', 'AVERAGE', 'AVERAGEX', 'COUNT', 'COUNTROWS', 'COUNTA', 'COUNTX',
                            'MIN', 'MINX', 'MAX', 'MAXX', 'DISTINCTCOUNT', 'DIVIDE',
                            'RELATED', 'RELATEDTABLE', 'USERELATIONSHIP', 'VALUES', 'DISTINCT',
                            'ADDCOLUMNS', 'SUMMARIZE', 'GROUPBY', 'EARLIER', 'EARLIEST', 'RANKX', 'TOPN',
                            'CALENDAR', 'CALENDARAUTO', 'DATE', 'TODAY', 'NOW', 'YEAR', 'MONTH', 'DAY',
                            'DATESYTD', 'DATEADD', 'SAMEPERIODLASTYEAR', 'TOTALYTD', 'PARALLELPERIOD',
                            'ISBLANK', 'IFERROR', 'HASONEVALUE', 'SELECTEDVALUE',
                            'FORMAT', 'CONCATENATE', 'CONCATENATEX', 'LEFT', 'RIGHT', 'MID', 'LEN',
                            'KEEPFILTERS', 'REMOVEFILTERS', 'TREATAS', 'CROSSFILTER',
                            'GENERATE', 'GENERATEALL', 'ROW', 'UNION', 'INTERSECT', 'EXCEPT',
                            'LOOKUPVALUE', 'SEARCH', 'FIND', 'CONTAINS'];
                        fn.forEach(function(f) {{
                            var re = new RegExp('\\\\b(' + f + ')\\\\s*\\\\(', 'gi');
                            esc = esc.replace(re, '###FUNCTION_START###$1###FUNCTION_END###(');
                        }});
                        
                        // Variáveis (após VAR)
                        esc = esc.replace(/(###KEYWORD_START###VAR###KEYWORD_END###)\\s+(\\w+)/gi, 
                            '$1 ###VARIABLE_START###$2###VARIABLE_END###');
                        
                        // Tabelas e Colunas
                        esc = esc.replace(/(\\w+)\\[([^\\]]+)\\]/g, 
                            '###TABLE_START###$1###TABLE_END###[###VARIABLE_START###$2###VARIABLE_END###]');
                        
                        // Converter marcadores para HTML (ÚLTIMO PASSO)
                        esc = esc.replace(/###COMMENT_START###/g, '<span class="dax-comment">');
                        esc = esc.replace(/###COMMENT_END###/g, '</span>');
                        esc = esc.replace(/###STRING_START###/g, '<span class="dax-string">');
                        esc = esc.replace(/###STRING_END###/g, '</span>');
                        esc = esc.replace(/###NUMBER_START###/g, '<span class="dax-number">');
                        esc = esc.replace(/###NUMBER_END###/g, '</span>');
                        esc = esc.replace(/###KEYWORD_START###/g, '<span class="dax-keyword">');
                        esc = esc.replace(/###KEYWORD_END###/g, '</span>');
                        esc = esc.replace(/###FUNCTION_START###/g, '<span class="dax-function">');
                        esc = esc.replace(/###FUNCTION_END###/g, '</span>');
                        esc = esc.replace(/###VARIABLE_START###/g, '<span class="dax-variable">');
                        esc = esc.replace(/###VARIABLE_END###/g, '</span>');
                        esc = esc.replace(/###TABLE_START###/g, '<span class="dax-table">');
                        esc = esc.replace(/###TABLE_END###/g, '</span>');
                        esc = esc.replace(/###OPERATOR_START###/g, '<span class="dax-operator">');
                        esc = esc.replace(/###OPERATOR_END###/g, '</span>');
                        
                        return esc;
                    }}

                    function resetGraphView() {{
                        try {{
                            if (typeof network !== 'undefined' && network) {{
                                network.fit({{animation: {{duration: 500, easingFunction: 'easeInOutQuad'}}}});
                            }} else {{
                                setTimeout(resetGraphView, 100);
                            }}
                        }} catch(e) {{
                            console.error('[DAX Viewer] Erro ao resetar:', e);
                        }}
                    }}

                    function colapsarRecursivo(noId) {{
                        var d = depsMap[noId];
                        if (d && d.filhos) {{
                            d.filhos.forEach(filhoId => {{
                                try {{
                                    var arestasConectadas = network.getConnectedEdges(filhoId);
                                    if (arestasConectadas.length <= 1) {{
                                        colapsarRecursivo(filhoId);
                                        nodes.remove(filhoId);
                                    }} else {{
                                        var arestaId = arestasConectadas.find(eId => {{
                                            var e = edges.get(eId);
                                            return (e.from == noId && e.to == filhoId) || (e.to == noId && e.from == filhoId);
                                        }});
                                        if (arestaId) edges.remove(arestaId);
                                    }}
                                }} catch(e) {{}}
                            }});
                        }}
                        var t = (infoData[noId] && infoData[noId].tipo) ? infoData[noId].tipo : "UNKNOWN";
                        var ic = iconesMap[t] || "❓";
                        var cr = coresMap[t] || "#CCCCCC";
                        nodes.update({{id: noId, label: ic + " " + noId + " ⊕", color: {{border: cr, background: cr}}, borderWidth: 3}});
                    }}

                    // Aguardar network estar pronto
                    function setupClickHandler() {{
                        if (typeof network === 'undefined') {{
                            console.log('[DAX Viewer] Aguardando network...');
                            setTimeout(setupClickHandler, 100);
                            return;
                        }}
                        
                        console.log('[DAX Viewer] Network pronto! Registrando evento de clique...');
                        
                        network.on("click", function(params) {{
                            console.log('[DAX Viewer] Click detectado!', params);
                            
                            if(params.nodes && params.nodes.length > 0) {{
                                var nodeId = params.nodes[0];
                                console.log('[DAX Viewer] Nó clicado:', nodeId);
                                
                                var nodeInfo = infoData[nodeId] || {{exp: 'Sem informação disponível', tipo: 'UNKNOWN'}};
                                console.log('[DAX Viewer] Info do nó:', nodeInfo);
                                
                                // Atualizar título
                                var titleEl = document.getElementById('p-title');
                                if (titleEl) {{
                                    titleEl.textContent = nodeId;
                                    console.log('[DAX Viewer] Título atualizado');
                                }} else {{
                                    console.error('[DAX Viewer] Elemento p-title não encontrado!');
                                }}
                                
                                // Atualizar código com highlighting
                                var expEl = document.getElementById('p-exp');
                                if (expEl) {{
                                    var highlighted = highlightDAX(nodeInfo.exp || 'Sem DAX');
                                    console.log('[DAX Viewer] HTML gerado:', highlighted.substring(0, 200));
                                    expEl.innerHTML = highlighted;
                                    console.log('[DAX Viewer] Código atualizado');
                                }} else {{
                                    console.error('[DAX Viewer] Elemento p-exp não encontrado!');
                                }}
                                
                                // Mostrar painel
                                var panel = document.getElementById('dax-panel');
                                if (panel) {{
                                    panel.style.display = 'block';
                                    console.log('[DAX Viewer] Painel exibido!');
                                }} else {{
                                    console.error('[DAX Viewer] Elemento dax-panel não encontrado!');
                                }}

                                // Lógica de expansão
                                if (modoExp) {{
                                    var d = depsMap[nodeId];
                                    if (d && d.filhos && d.filhos.length) {{
                                        var jaExpandido = false;
                                        try {{
                                            var connectedNodes = network.getConnectedNodes(nodeId);
                                            jaExpandido = d.filhos.some(fId => connectedNodes.includes(fId));
                                        }} catch(e) {{}}

                                        if (jaExpandido) {{
                                            colapsarRecursivo(nodeId);
                                        }} else {{
                                            var icPrincipal = iconesMap[nodeInfo.tipo] || "❓";
                                            nodes.update({{id: nodeId, label: icPrincipal + " " + nodeId + " ⊖", color: {{border: "#FF4B4B"}}, borderWidth: 4}});
                                            d.filhos.forEach((f, idx) => {{
                                                var t_f = d.tipos[idx];
                                                var ic_f = iconesMap[t_f] || "❓";
                                                var cr_f = coresMap[t_f] || "#CCCCCC";
                                                try {{
                                                    if (!nodes.get(f)) {{
                                                        var temFilhos = (depsMap[f] && depsMap[f].filhos && depsMap[f].filhos.length > 0);
                                                        nodes.add({{id: f, label: ic_f + " " + f + (temFilhos ? " ⊕" : ""), color: cr_f, shape: "box", font: {{face: "Segoe UI", size: 14, bold: temFilhos}}, borderWidth: temFilhos ? 3 : 1}});
                                                    }}
                                                    edges.add({{from: nodeId, to: f, color: "#CCCCCC", width: 1}});
                                                }} catch(e) {{}}
                                            }});
                                        }}
                                    }}
                                }}
                            }} else {{
                                console.log('[DAX Viewer] Nenhum nó selecionado');
                            }}
                        }});
                        
                        console.log('[DAX Viewer] Evento registrado com sucesso!');
                    }}
                    
                    // Iniciar configuração
                    setupClickHandler();
                </script>
                """
                
                # Injetar HTML no corpo do grafo
                html_final = h_base.replace("</body>", f"{estilos_css}{painel_html}{script_js}</body>")
                
                st.subheader("Visualização do Grafo")
                components.html(html_final, height=650)

                # 6. Exportação Sidebar
                with export_placeholder:
                    st.sidebar.markdown("---")
                    st.sidebar.download_button("📸 Exportar Grafo (HTML)", html_final, "grafo.html", "text/html", type="primary", use_container_width=True)

                # 7. Análise de Impacto Detalhada
                st.markdown("---")
                st.subheader("Análise de Impacto por Medida")
                cols = st.columns(min(3, len(medidas_selecionadas)))
                for idx, m_name in enumerate(medidas_selecionadas):
                    with cols[idx % 3]:
                        if m_name in G:
                            with st.container(border=True):
                                st.markdown(f"**{m_name}**")
                                desc = nx.descendants(G, m_name)
                                if desc:
                                    with st.expander(f"Dependências ({len(desc)})"): st.write(sorted(list(desc)))
                                
                                # USO EM PÁGINAS (corrigido para usar mesma lógica da métrica)
                                if df_st is not None:
                                    paginas_uso = []
                                    for _, row_pg in df_st.iterrows():
                                        meds_raw_pg = str(row_pg['Medidas'])
                                        if meds_raw_pg and meds_raw_pg != 'nan' and meds_raw_pg != 'None':
                                            meds_in_pg = [x.strip() for x in meds_raw_pg.split(',') if x.strip()]
                                            if m_name in meds_in_pg:
                                                paginas_uso.append(row_pg['Página'])
                                    
                                    paginas_uso = sorted(list(set(paginas_uso)))  # Remove duplicatas e ordena
                                    st.markdown(f"📄 **Uso em Páginas:** {len(paginas_uso)}")
                                    if len(paginas_uso) > 0:
                                        with st.expander("Ver lista de páginas"):
                                            for p_name in paginas_uso:
                                                st.markdown(f"✅ `{p_name}`")
                                    else:
                                        st.caption("Esta medida não foi encontrada em nenhum visual de página.")

    else:
        st.error("Colunas [Origem] ou [Destino] não encontradas no arquivo.")
else:
    st.info("Aguardando upload do arquivo para gerar o dashboard.")